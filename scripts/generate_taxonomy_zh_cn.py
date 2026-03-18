#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "Multiling IOC 15.1_d.xlsx"
LABELS_PATH = ROOT / "analyzer" / "models" / "labels.txt"
OUT_PATH = ROOT / "analyzer" / "taxonomy_zh_cn.json"


MANUAL_IOC_ALIASES = {
    "Bank Swallow": "Sand Martin",
    "Barn Owl": "American Barn Owl",
    "Black Swift": "American Black Swift",
    "Black-bellied Plover": "Grey Plover",
    "Brant": "Brant Goose",
    "Bushtit": "American Bushtit",
    "Cattle Egret": "Western Cattle Egret",
    "Chukar": "Chukar Partridge",
    "Cliff Swallow": "American Cliff Swallow",
    "Common Raven": "Northern Raven",
    "Common Redpoll": "Redpoll",
    "Dovekie": "Little Auk",
    "Dusky Flycatcher": "American Dusky Flycatcher",
    "Eared Grebe": "Black-necked Grebe",
    "European Starling": "Common Starling",
    "Fox Sparrow": "Red Fox Sparrow",
    "Herring Gull": "American Herring Gull",
    "Hoary Redpoll": "Redpoll",
    "House Wren": "Northern House Wren",
    "Northern Goshawk": "American Goshawk",
    "Pacific-slope Flycatcher": "Western Flycatcher",
    "Ring-necked Pheasant": "Common Pheasant",
    "Rock Pigeon": "Rock Dove",
    "Rough-legged Hawk": "Rough-legged Buzzard",
    "Whimbrel": "Hudsonian Whimbrel",
    "White Ibis": "American White Ibis",
    "White-winged Crossbill": "Two-barred Crossbill",
    "Yellow Warbler": "American Yellow Warbler",
}

MANUAL_FIXED_TRANSLATIONS = {
    # IOC 15.1 no longer exposes this legacy project label as a direct English row.
    "Yellow-rumped Warbler": "黄腰林莺",
}


def normalize_label(text: str) -> str:
    normalized = text.lower().strip()
    normalized = normalized.replace("gray", "grey")
    normalized = normalized.replace("'", "")
    normalized = re.sub(r"[^a-z0-9]+", "", normalized)
    return normalized


def load_labels(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig") as f:
        return [line.strip() for line in f if line.strip()]


def load_ioc_species(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["List"]
    species: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[4]:
            continue
        english = str(row[4]).strip()
        species[english] = {
            "scientific_name": str(row[3] or "").strip(),
            "scientific_family": str(row[2] or "").strip(),
            "zh_cn": str(row[6] or "").strip(),
            "zh_tw": str(row[7] or "").strip(),
        }
    return species


def build_species_map(project_labels: list[str], ioc_species: dict[str, dict[str, str]]) -> dict:
    normalized_ioc: dict[str, str] = {}
    for english in ioc_species:
        key = normalize_label(english)
        if key not in normalized_ioc:
            normalized_ioc[key] = english

    species_map: dict[str, str] = {}
    species_ioc_english: dict[str, str] = {}
    match_types = {
        "direct": 0,
        "normalized": 0,
        "manual_alias": 0,
        "manual_fixed": 0,
    }
    unresolved: list[str] = []

    for label in project_labels:
        if label in ioc_species:
            species_map[label] = ioc_species[label]["zh_cn"]
            species_ioc_english[label] = label
            match_types["direct"] += 1
            continue

        normalized = normalize_label(label)
        normalized_match = normalized_ioc.get(normalized)
        if normalized_match:
            species_map[label] = ioc_species[normalized_match]["zh_cn"]
            species_ioc_english[label] = normalized_match
            match_types["normalized"] += 1
            continue

        alias = MANUAL_IOC_ALIASES.get(label)
        if alias:
            species_map[label] = ioc_species[alias]["zh_cn"]
            species_ioc_english[label] = alias
            match_types["manual_alias"] += 1
            continue

        fixed = MANUAL_FIXED_TRANSLATIONS.get(label)
        if fixed:
            species_map[label] = fixed
            species_ioc_english[label] = ""
            match_types["manual_fixed"] += 1
            continue

        unresolved.append(label)

    if unresolved:
        raise RuntimeError(f"Unresolved species labels: {unresolved}")

    return {
        "species": species_map,
        "species_ioc_english": species_ioc_english,
        "match_types": match_types,
    }


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {WORKBOOK_PATH}")

    project_labels = load_labels(LABELS_PATH)
    ioc_species = load_ioc_species(WORKBOOK_PATH)
    species_result = build_species_map(project_labels, ioc_species)

    payload = {
        "meta": {
            "generated_at_utc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "source": {
                "workbook": WORKBOOK_PATH.name,
                "sheet": "List",
                "english_column": "English",
                "zh_cn_column": "Chinese",
            },
            "stats": {
                "project_species_labels": len(project_labels),
                "ioc_species_rows": len(ioc_species),
                **species_result["match_types"],
            },
            "notes": [
                "Species labels come from the official IOC multilingual workbook.",
                "Some project labels require normalization or legacy alias mapping because the model uses older English common names.",
                "Family Chinese names are intentionally left empty until an authoritative family-level source is added.",
            ],
        },
        "species": species_result["species"],
        "species_ioc_english": species_result["species_ioc_english"],
        "family_display": {},
        "family_scientific": {},
    }

    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    stats = payload["meta"]["stats"]
    print(f"Wrote {OUT_PATH}")
    print(
        "Species coverage:",
        f"{stats['project_species_labels']} labels",
        f"(direct={stats['direct']}, normalized={stats['normalized']},",
        f"manual_alias={stats['manual_alias']}, manual_fixed={stats['manual_fixed']})",
    )


if __name__ == "__main__":
    main()
